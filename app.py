from flask import Flask, render_template, request, redirect, url_for, send_file, session
from services.vt_service import check_virustotal
from services.abuse_service import check_abuseipdb
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
#from services.hetrixtools_service import check_hetrixtools
#from services.talos_service import check_talos
import time
import re

app = Flask(__name__)
app.secret_key = "clave_super_secreta"


def detectar_tipo(indicador):

    indicador = indicador.strip()

    if re.match(r"^\d{1,3}(\.\d{1,3}){3}$", indicador):
        return "ip"

    if indicador.startswith("http://") or indicador.startswith("https://"):
        return "url"

    if re.match(r"^[a-fA-F0-9]{32}$", indicador) \
       or re.match(r"^[a-fA-F0-9]{40}$", indicador) \
       or re.match(r"^[a-fA-F0-9]{64}$", indicador):
        return "hash"

    if "." in indicador:
        return "domain"

    return "unknown"


@app.route("/", methods=["GET", "POST"])
def index():

    if request.method == "POST":

        indicadores = request.form.get("indicadores", "").splitlines()
        results = []
        start_time = time.time()

        for indicador in indicadores:
            indicador = indicador.strip()
            if not indicador:
                continue

            tipo = detectar_tipo(indicador)
            #print("TIPO DETECTADO:", tipo, "INDICADOR:", indicador)
            vt_data = check_virustotal(indicador, tipo)

            if tipo == "ip":
                abuse_data = check_abuseipdb(indicador)
                # hetrix_data = check_hetrixtools(indicador)
                #talos_data = check_talos(indicador)
                #print("DEBUG TALOS:", talos_data)
                #talos_reputation = talos_data.get("talos_reputation", "-")
                abuse_score = abuse_data.get("abuse_score", 0)
                country = abuse_data.get("country", "Desconocido")
               # hetrix_blacklists = hetrix_data.get("blacklists", 0)
                #talos_reputation = talos_data.get("talos_reputation", "-")
            else:
                abuse_score = "-"
                country = "-"
                # hetrix_blacklists = "-"
                #talos_reputation = "-"

            results.append({
                "indicador": indicador,
                "tipo": tipo.upper(),
                "vt_malicious": vt_data["vt_malicious"],
                "vt_suspicious": vt_data["vt_suspicious"],
                #"hetrix_blacklists": hetrix_blacklists,
                #"talos_reputation": talos_reputation,
                "abuse_score": abuse_score,
                "country": country
            })

        elapsed_time = round(time.time() - start_time, 2)

        session["results"] = results
        session["elapsed_time"] = elapsed_time

        return redirect(url_for("index"))

    # 🔥 IMPORTANTE: usamos pop para limpiar al recargar
    results = session.get("results")
    elapsed_time = session.get("elapsed_time")

    return render_template("index.html",
                           results=results,
                           elapsed_time=elapsed_time)


@app.route("/descargar")
def descargar():

    results = session.get("results")
    if not results:
        return redirect(url_for("index"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = [
    "Indicador",
    "Tipo Detectado",
    "Estado",
    "Número de reportes maliciosos en VirusTotal",
    "Reportes sospechosos en VirusTotal",
    #"Reportes HetrixTools (Blacklists)"
    "Puntaje de AbuseIPDB (%)",
    "País"
]

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    for row in results:

        if row["tipo"] == "IP":
            estado = "Malicioso" if row["vt_malicious"] > 0 or row["abuse_score"] > 50 else "No malicioso"
            abuse_value = row["abuse_score"] / 100
        else:
            estado = "Malicioso" if row["vt_malicious"] > 0 else "No malicioso"
            abuse_value = "-"

        ws.append([
            row["indicador"],
            row["tipo"],
            estado,
            row["vt_malicious"],
            row["vt_suspicious"],
            #row["hetrix_blacklists"],
            abuse_value,
            row["country"]
        ])

    for r in range(2, ws.max_row + 1):
        if ws[f"F{r}"].value != "-":
            ws[f"F{r}"].number_format = "0%"

    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file,
                     download_name="reporte_ti.xlsx",
                     as_attachment=True)

@app.route("/limpiar")
def limpiar():
    session.pop("results", None)
    session.pop("elapsed_time", None)
    return redirect(url_for("index"))

if __name__ == "__main__":
    app.run(debug=True)